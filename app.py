import io
import re
import zipfile
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell


# =========================
# Konstanta header (Mass Update)
# =========================
MASS_HEADER_SKU = "SKU Penjual"
MASS_HEADER_PRICE = "Harga Ritel (Mata Uang Lokal)"

# Pricelist: header minimal yang kita cari
PL_HEADER_SKU_CANDIDATES = ["KODEBARANG", "KODE BARANG", "SKU", "SKU NO", "SKU_NO", "KODEBARANG "]
PL_PRICE_COL_TIKTOK = "M3"
PL_PRICE_COL_SHOPEE = "M4"

# Addon mapping: header yang diharapkan (boleh lebih dari 1 kandidat)
ADDON_CODE_CANDIDATES = ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"]
ADDON_PRICE_CANDIDATES = ["harga", "HARGA", "Price", "PRICE", "Harga"]

# Heuristik: jika nilai < 1.000.000 dianggap "tanpa 000" dan perlu dikali 1000
SMALL_TO_THOUSAND_THRESHOLD = 1_000_000
AUTO_MULTIPLIER_FOR_SMALL = 1000

SKU_SPLIT_PLUS = re.compile(r"\+")


# =========================
# Utils
# =========================
def normalize_text(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def normalize_addon_code(x) -> str:
    # addon code tidak peka besar/kecil
    return normalize_text(x).upper()


def detect_platform_from_filename(filename: str) -> str:
    """Return 'tiktok' or 'shopee'. Default: tiktok (M3)."""
    name = (filename or "").lower()
    if "shopee" in name:
        return "shopee"
    if "tiktok" in name:
        return "tiktok"
    # default
    return "tiktok"


def parse_platform_sku(full_sku: str) -> Tuple[str, List[str]]:
    """
    Contoh: ND-LAP-LE-82XQ00HX1D+PC+BA
    base = ND-LAP-LE-82XQ00HX1D
    addons = ["PC","BA"]
    """
    if full_sku is None:
        return "", []

    s = str(full_sku).strip()
    if not s:
        return "", []

    parts = SKU_SPLIT_PLUS.split(s)
    base = parts[0].strip()
    addons = [p.strip() for p in parts[1:] if p and p.strip()]
    return base, addons


def parse_number_like_id(x) -> str:
    """
    Biar SKU yang kebaca angka nggak jadi 1.234E+12.
    Kalau x numeric -> convert ke int string (tanpa .0).
    """
    if x is None:
        return ""
    if isinstance(x, (int,)):
        return str(x)
    if isinstance(x, float):
        if x.is_integer():
            return str(int(x))
        return str(x)
    return str(x).strip()


def parse_price_cell(val) -> Optional[int]:
    """
    Return integer Rupiah (tanpa simbol).
    Bisa handle:
      - 9300 -> dianggap 9.300.000 (akan dikali 1000 oleh apply_multiplier_if_needed)
      - "9.300" -> jadi 9300
      - "15,900" -> jadi 15.9 (jarang) -> kita handle secara aman
    """
    if val is None:
        return None

    # Excel numeric
    if isinstance(val, (int, float)):
        try:
            if isinstance(val, float) and val.is_integer():
                return int(val)
            return int(round(float(val)))
        except Exception:
            return None

    s = str(val).strip()
    if not s:
        return None

    # buang spasi & mata uang
    s = s.replace("Rp", "").replace("rp", "").replace(" ", "")

    # kasus indonesia: 9.300 (titik ribuan)
    # kalau ada koma sebagai desimal, kita ubah dulu
    # strategi:
    #  - jika ada '.' dan tidak ada ',' -> remove '.'
    #  - jika ada ',' dan '.' -> asumsi '.' ribuan, ',' desimal -> remove '.', replace ',' with '.'
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s and "," not in s:
        s = s.replace(".", "")
    elif "," in s and "." not in s:
        # anggap koma pemisah ribuan? (jarang) -> remove ','
        s = s.replace(",", "")

    # keep digits + optional decimal
    try:
        f = float(s)
        if f.is_integer():
            return int(f)
        return int(round(f))
    except Exception:
        return None


def apply_multiplier_if_needed(x: int) -> int:
    """
    Kalau angka kecil (<1jt), kita anggap itu 'tanpa 000' -> x*1000.
    Kalau sudah besar, biarkan.
    """
    if x is None:
        return 0
    if x < SMALL_TO_THOUSAND_THRESHOLD:
        return x * AUTO_MULTIPLIER_FOR_SMALL
    return x


def safe_set_cell_value(ws, row: int, col: int, value):
    """
    Aman untuk merged cell: kalau cell target adalah MergedCell (read-only),
    kita tulis ke top-left dari merged range.
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # cari merged range yang mencakup cell ini
        coord = cell.coordinate
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                # tulis ke start cell
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return
        # fallback: skip
        return
    cell.value = value


# =========================
# Excel scanning: find header row & column indexes
# =========================
def find_header_row_and_cols_mass(ws) -> Tuple[int, int, int]:
    """
    Cari header row yang mengandung:
      - "SKU Penjual"
      - "Harga Ritel (Mata Uang Lokal)"
    Return: (header_row_idx, sku_col_idx, price_col_idx)
    """
    target_a = MASS_HEADER_SKU.strip().lower()
    target_b = MASS_HEADER_PRICE.strip().lower()

    for r in range(1, 30):  # scan 1..29
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row_vals.append("")
            else:
                row_vals.append(str(v).strip())

        # mapping lower->col
        lower_map = {str(v).strip().lower(): i + 1 for i, v in enumerate(row_vals) if str(v).strip() != ""}
        if target_a in lower_map and target_b in lower_map:
            return r, lower_map[target_a], lower_map[target_b]

    raise ValueError(f"Header Mass Update tidak ketemu. Pastikan ada '{MASS_HEADER_SKU}' dan '{MASS_HEADER_PRICE}'.")


def find_header_row_and_cols_pricelist(ws) -> Tuple[int, int, int, int]:
    """
    Cari header row yang mengandung SKU/KODEBARANG dan M3 & M4.
    Return: (header_row_idx, sku_col_idx, m3_col_idx, m4_col_idx)
    """
    candidates = [c.strip().lower() for c in PL_HEADER_SKU_CANDIDATES]
    target_m3 = PL_PRICE_COL_TIKTOK.lower()
    target_m4 = PL_PRICE_COL_SHOPEE.lower()

    for r in range(1, 60):  # pricelist biasanya banyak header
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append("" if v is None else str(v).strip())

        lower_to_col = {}
        for idx, v in enumerate(row_vals, start=1):
            lv = v.strip().lower()
            if not lv:
                continue
            # jika duplicate header, ambil yang pertama saja
            if lv not in lower_to_col:
                lower_to_col[lv] = idx

        sku_col = None
        for cand in candidates:
            if cand in lower_to_col:
                sku_col = lower_to_col[cand]
                break

        if sku_col is not None and target_m3 in lower_to_col and target_m4 in lower_to_col:
            return r, sku_col, lower_to_col[target_m3], lower_to_col[target_m4]

    raise ValueError("Header Pricelist tidak ketemu. Pastikan ada kolom KODEBARANG (atau setara) dan kolom M3 & M4.")


def load_pricelist_map(pl_bytes: bytes) -> Dict[str, Dict[str, int]]:
    """
    Return map:
      base_sku -> {"M3": price_int, "M4": price_int}
    Harga akan auto *1000 jika angkanya kecil.
    """
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True)
    ws = wb.active

    header_row, sku_col, m3_col, m4_col = find_header_row_and_cols_pricelist(ws)

    m: Dict[str, Dict[str, int]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sku_val = ws.cell(row=r, column=sku_col).value
        sku = normalize_text(sku_val)
        if not sku:
            continue

        # normalize sku for matching: keep as-is
        m3_raw = parse_price_cell(ws.cell(row=r, column=m3_col).value)
        m4_raw = parse_price_cell(ws.cell(row=r, column=m4_col).value)

        if m3_raw is None and m4_raw is None:
            continue

        m3 = apply_multiplier_if_needed(m3_raw) if m3_raw is not None else None
        m4 = apply_multiplier_if_needed(m4_raw) if m4_raw is not None else None

        m[sku] = {}
        if m3 is not None:
            m[sku]["M3"] = int(m3)
        if m4 is not None:
            m[sku]["M4"] = int(m4)

    return m


def load_addon_map(addon_bytes: bytes) -> Dict[str, int]:
    """
    Addon mapping Excel:
      kolom addon_code / harga (nama bisa beda), kita cari dari kandidat.
    Harga juga auto *1000 jika kecil.
    Return:
      ADDON_CODE_UPPER -> harga_int
    """
    wb = load_workbook(io.BytesIO(addon_bytes), data_only=True)
    ws = wb.active

    # cari header row dalam 1..30
    header_row = None
    code_col = None
    price_col = None

    code_cands = [c.strip().lower() for c in ADDON_CODE_CANDIDATES]
    price_cands = [c.strip().lower() for c in ADDON_PRICE_CANDIDATES]

    for r in range(1, 30):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append("" if v is None else str(v).strip())

        lower_to_col = {}
        for idx, v in enumerate(row_vals, start=1):
            lv = v.strip().lower()
            if not lv:
                continue
            if lv not in lower_to_col:
                lower_to_col[lv] = idx

        found_code = None
        for cc in code_cands:
            if cc in lower_to_col:
                found_code = lower_to_col[cc]
                break

        found_price = None
        for pc in price_cands:
            if pc in lower_to_col:
                found_price = lower_to_col[pc]
                break

        if found_code and found_price:
            header_row = r
            code_col = found_code
            price_col = found_price
            break

    if header_row is None or code_col is None or price_col is None:
        raise ValueError("Header Addon Mapping tidak ketemu. Pastikan ada kolom addon_code & harga (atau setara).")

    m: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = normalize_addon_code(ws.cell(row=r, column=code_col).value)
        if not code:
            continue

        price_raw = parse_price_cell(ws.cell(row=r, column=price_col).value)
        if price_raw is None:
            continue

        price = apply_multiplier_if_needed(int(price_raw))
        m[code] = int(price)

    return m


@dataclass
class RowChange:
    file: str
    excel_row: int
    sku_full: str
    old_price: int
    new_price: int
    reason: str


def compute_new_price_for_row(
    sku_full: str,
    platform: str,
    pl_map: Dict[str, Dict[str, int]],
    addon_map: Dict[str, int],
    discount_rp: int,
) -> Tuple[Optional[int], str]:
    """
    Return (new_price or None, reason).
    If None -> do not change anything.
    Rules:
      - base SKU must exist in pricelist with M3/M4 based on platform
      - if ANY addon code missing in addon_map -> None (do not change)
      - final = base + sum(addons) - discount_rp
      - no rounding
    """
    base_sku, addons = parse_platform_sku(sku_full)
    if not base_sku:
        return None, "SKU kosong"

    pl = pl_map.get(base_sku)
    if not pl:
        return None, "Base SKU tidak ada di Pricelist"

    price_key = "M3" if platform == "tiktok" else "M4"
    base_price = pl.get(price_key)
    if base_price is None:
        return None, f"Harga {price_key} kosong di Pricelist"

    # addon: case-insensitive
    addon_total = 0
    for a in addons:
        code = normalize_addon_code(a)
        if not code:
            continue
        if code not in addon_map:
            return None, f"Addon '{code}' tidak ada di file Addon Mapping"
        addon_total += int(addon_map[code])

    final_price = int(base_price) + int(addon_total) - int(discount_rp)
    if final_price < 0:
        final_price = 0

    return final_price, f"{price_key} + addon - diskon"


def make_issues_workbook(changes: List[RowChange]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "issues_report"
    headers = ["file", "row", "sku_full", "old_price", "new_price", "reason"]
    ws.append(headers)
    for ch in changes:
        ws.append([ch.file, ch.excel_row, ch.sku_full, ch.old_price, ch.new_price, ch.reason])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def workbook_to_bytes(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# UI
# =========================
st.set_page_config(page_title="Web App Update Harga", layout="wide")
st.title("Web App Update Harga")

# Upload row: 3 columns (bagusan yang sejajar)
c1, c2, c3 = st.columns(3)
with c1:
    mass_files = st.file_uploader(
        "Upload Mass Update (bisa banyak)",
        type=["xlsx"],
        accept_multiple_files=True,
    )
with c2:
    pl_file = st.file_uploader("Upload Pricelist", type=["xlsx"])
with c3:
    addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"])

st.divider()

discount_rp = st.number_input("Diskon (Rp) - mengurangi harga final", min_value=0, value=0, step=1000)

process = st.button("Proses")

if process:
    if not mass_files or pl_file is None or addon_file is None:
        st.error("Wajib upload: Mass Update (minimal 1), Pricelist, dan Addon Mapping.")
        st.stop()

    try:
        pl_map = load_pricelist_map(pl_file.getvalue())
    except Exception as e:
        st.error(f"Gagal baca Pricelist: {e}")
        st.stop()

    try:
        addon_map = load_addon_map(addon_file.getvalue())
    except Exception as e:
        st.error(f"Gagal baca Addon Mapping: {e}")
        st.stop()

    changed_rows: List[RowChange] = []
    output_files: List[Tuple[str, bytes]] = []

    for mf in mass_files:
        filename = mf.name
        platform = detect_platform_from_filename(filename)

        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active

        try:
            header_row, sku_col, price_col = find_header_row_and_cols_mass(ws)
        except Exception as e:
            changed_rows.append(RowChange(
                file=filename,
                excel_row=0,
                sku_full="",
                old_price=0,
                new_price=0,
                reason=f"Gagal baca header mass update: {e}",
            ))
            continue

        # iterasi data mulai header_row+1 sampai max_row
        # (data kamu biasanya mulai baris 6, tapi kita biarkan auto dari header)
        file_changed_count = 0
        for r in range(header_row + 1, ws.max_row + 1):
            sku_val = ws.cell(row=r, column=sku_col).value
            sku_full = parse_number_like_id(sku_val)
            if not sku_full:
                # skip baris kosong, tapi jangan break (kadang ada blank di tengah)
                continue

            old_price_raw = parse_price_cell(ws.cell(row=r, column=price_col).value)
            old_price = int(old_price_raw) if old_price_raw is not None else 0

            new_price, reason = compute_new_price_for_row(
                sku_full=sku_full,
                platform=platform,
                pl_map=pl_map,
                addon_map=addon_map,
                discount_rp=int(discount_rp),
            )

            if new_price is None:
                # aturan kamu: kalau ada 1 saja yang gak ketemu -> tidak ubah apapun
                continue

            if int(new_price) == int(old_price):
                continue

            # set value aman
            safe_set_cell_value(ws, row=r, col=price_col, value=int(new_price))
            file_changed_count += 1

            changed_rows.append(RowChange(
                file=filename,
                excel_row=r,
                sku_full=sku_full,
                old_price=int(old_price),
                new_price=int(new_price),
                reason=reason,
            ))

        out_bytes = workbook_to_bytes(wb)
        out_name = filename.replace(".xlsx", "_updated.xlsx")
        output_files.append((out_name, out_bytes))

    # Preview: hanya yang berubah
    st.subheader("Preview (yang berubah saja)")
    if not changed_rows:
        st.info("Tidak ada perubahan harga (mungkin SKU tidak ketemu di Pricelist / addon tidak cocok / atau harga sama).")
    else:
        # tampilkan 200 pertama biar ringan
        import pandas as pd
        df_preview = pd.DataFrame([{
            "file": x.file,
            "row": x.excel_row,
            "sku_full": x.sku_full,
            "old_price": x.old_price,
            "new_price": x.new_price,
            "reason": x.reason,
        } for x in changed_rows])
        st.dataframe(df_preview.head(200), use_container_width=True)

    st.divider()

    # Download output: kalau banyak -> ZIP
    if len(output_files) == 1:
        name, data = output_files[0]
        st.download_button(
            "Download hasil (XLSX)",
            data=data,
            file_name=name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        zbuf = io.BytesIO()
        with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, data in output_files:
                zf.writestr(name, data)
            # juga masukkan report perubahan
            rep = make_issues_workbook(changed_rows)
            zf.writestr("changes_report.xlsx", rep)

        st.download_button(
            "Download semua hasil (ZIP)",
            data=zbuf.getvalue(),
            file_name="mass_update_results.zip",
            mime="application/zip",
        )

    # Report perubahan juga terpisah
    if changed_rows:
        rep_bytes = make_issues_workbook(changed_rows)
        st.download_button(
            "Download Report Perubahan (XLSX)",
            data=rep_bytes,
            file_name="changes_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )